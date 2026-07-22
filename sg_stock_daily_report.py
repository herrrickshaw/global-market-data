# sg_stock_daily_report.py
# ==========================
# Daily stock report for Singapore-listed equities (SGX).
# All data is sourced from Yahoo Finance via the yfinance library.
#
# Sections per stock:
#   • Live quote, valuation ratios (CMP, P/E, P/B)
#   • Corporate actions (dividends, splits)
#   • 30-day historical price summary
#
# Quantitative scans:
#   • Darvas Box       — technical momentum breakout
#   • Piotroski F-Score — 9-point financial-strength score
#   • PEGY Ratio       — PEG adjusted for earnings growth + dividend yield
#   • Breakout Opportunities — price vs 200-day MA
#
# Install dependencies (run once):
#   pip install yfinance pandas openpyxl lxml
#
# Single-stock usage:
#   python sg_stock_daily_report.py D05
#   python sg_stock_daily_report.py BN4 --scans
#   python sg_stock_daily_report.py U11 --output excel
#
# Batch usage:
#   python sg_stock_daily_report.py --sti30
#   python sg_stock_daily_report.py --sti30 --scans --output excel
#
# Colab quick-start:
#   !pip install yfinance pandas openpyxl lxml
#   from sg_stock_daily_report import run, run_batch, run_scans_only
#   run("D05", run_scans=True)
#   run_batch(run_scans=True, output_format="excel")

# ── Standard library ──────────────────────────────────────────────────────────
import argparse
import sys
from datetime import datetime
from pathlib import Path

# ── Third-party ───────────────────────────────────────────────────────────────
import pandas as pd

try:
    import yfinance as yf
except ImportError:
    sys.exit("❌  pip install yfinance")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    sys.exit("❌  pip install openpyxl")


# ── Constants ─────────────────────────────────────────────────────────────────
DOWNLOAD_DIR = Path("./sg_stock_data")
DOWNLOAD_DIR.mkdir(exist_ok=True)

DARVAS_CONFIRM = 3  # consecutive days a high/low must hold to confirm a box

# Straits Times Index (STI) — 30 largest Singapore-listed companies
STI_30 = [
    "D05.SI",
    "BN4.SI",
    "U11.SI",
    "C6L.SI",
    "C38U.SI",
    "ME8U.SI",
    "N2IU.SI",
    "S58.SI",
    "Z74.SI",
    "TSM.SI",
    "G13.SI",
    "5ZB.SI",
    "J36.SI",
    "A17U.SI",
    "C52.SI",
    "T82U.SI",
    "M44U.SI",
    "BS6.SI",
    "S63.SI",
    "AWX.SI",
    "E27.SI",
    "TS0U.SI",
    "AJBU.SI",
    "S51.SI",
    "VOID.SI",
    "F3L.SI",
    "CCHT.SI",
    "9CI.SI",
    "U96.SI",
    "S58.SI",
]

# Alternative: Top 15 by market cap
STI_TOP_15 = [
    "D05.SI",  # DBS Group
    "BN4.SI",  # UOB
    "U11.SI",  # Singtel
    "C6L.SI",  # Keppel
    "C38U.SI",  # Mapletree Commercial Trust
    "ME8U.SI",  # Mapletree Industrial
    "N2IU.SI",  # Ascendas REIT
    "S58.SI",  # Singapore Airlines
    "Z74.SI",  # Stanchart
    "TSM.SI",  # Thai Beverage
    "G13.SI",  # Genting Singapore
    "5ZB.SI",  # Seatrium
    "J36.SI",  # Jiutanji Holdings
    "A17U.SI",  # Ascendas India
    "C52.SI",  # Frencken Group
]

# Complete SGX Stock List (175+ active stocks, 2026)
SGX_ALL = [
    # Main Board Blue Chips (STI Components)
    "D05.SI",
    "BN4.SI",
    "U11.SI",
    "C6L.SI",
    "C38U.SI",
    "ME8U.SI",
    "N2IU.SI",
    "S58.SI",
    "Z74.SI",
    "TSM.SI",
    "G13.SI",
    "5ZB.SI",
    "J36.SI",
    "A17U.SI",
    "C52.SI",
    "T82U.SI",
    "M44U.SI",
    "BS6.SI",
    "S63.SI",
    "AWX.SI",
    "E27.SI",
    "TS0U.SI",
    "AJBU.SI",
    "S51.SI",
    "VOID.SI",
    "F3L.SI",
    "CCHT.SI",
    "9CI.SI",
    "U96.SI",
    # Major REITs
    "AV8U.SI",
    "BUOU.SI",
    "CLR.SI",
    "CMOU.SI",
    # Banks & Finance
    "O39.SI",
    "O5RU.SI",
    # Infrastructure & Utilities
    "U14.SI",
    "F01.SI",
    "CLI.SI",
    "LMW.SI",
    # Technology & Semiconductors
    "TL0.SI",
    "KURO.SI",
    "JAC.SI",
    "M1GU.SI",
    # Manufacturing & Engineering
    "APL.SI",
    "LHN.SI",
    "PMI.SI",
    "PTNK.SI",
    # Retail & Consumer
    "BDG.SI",
    "F34.SI",
    "WSH.SI",
    "CFFP.SI",
    # Marine & Shipping
    "AIS.SI",
    "KAMU.SI",
    "ZIP.SI",
    "SWIMC.SI",
    # Oil & Gas
    "ES3.SI",
    "JHQU.SI",
    # Trading & Distribution
    "5LA.SI",
    "TIW.SI",
    "VFG.SI",
    # Properties & Real Estate
    "OXAM.SI",
    "T60U.SI",
    "SPDH.SI",
    # Healthcare
    "M01.SI",
    "MLP.SI",
    # Hospitality & Gaming
    "MSLH.SI",
    # Fashion & Retail
    "ACV.SI",
    "CC3U.SI",
    # Food & Beverage
    "FJM.SI",
    # Biotech & Pharma
    "LFS.SI",
    "IHH.SI",
    # Additional Active Stocks (Cyclicals, SMEs, Growth)
    "CHEU.SI",
    "CLRC.SI",
    "CMS.SI",
    "CRPH.SI",
    "CSE.SI",
    "CTT.SI",
    "D01.SI",
    "D04.SI",
    "D11.SI",
    "D22.SI",
    "D24.SI",
    "D28.SI",
    "D29.SI",
    "D32.SI",
    "D33.SI",
    "DCS.SI",
    "DGS.SI",
    "DHI.SI",
    "DJ5.SI",
    "DPL.SI",
    "DQ5.SI",
    "DRL.SI",
    "DRX.SI",
    "DTY.SI",
    "DUL.SI",
    "DXQ.SI",
    "E2F.SI",
    "E5F.SI",
    "EL1.SI",
    "ELD.SI",
    "ELM.SI",
    "EQ5.SI",
    "ESPL.SI",
    "EVT.SI",
    "F2G.SI",
    "F6G.SI",
    "F8K.SI",
    "FA8.SI",
    "FAX.SI",
    "FEL.SI",
    "FEV.SI",
    "FGL.SI",
    "FH1.SI",
    "FHS.SI",
    "FJL.SI",
    "FKA.SI",
    "FKG.SI",
    "FLA.SI",
    "FLE.SI",
    "FLJ.SI",
    "FMF.SI",
    "FOL.SI",
    "FPC.SI",
    "FPT.SI",
    "FPV.SI",
    "FPW.SI",
    "FPY.SI",
    "FRD.SI",
    "FRP.SI",
    "FS1.SI",
    "FSD.SI",
    "FSE.SI",
    "FSL.SI",
    "FSR.SI",
    "FSU.SI",
    "FVK.SI",
    "FWL.SI",
    "FY8.SI",
    "FZE.SI",
    "G08.SI",
    "G1G.SI",
    "G4Q.SI",
    "G92.SI",
    "G93.SI",
    "G94.SI",
    "G95.SI",
    "G96.SI",
    "G97.SI",
    "G98.SI",
    "GA8.SI",
    "GAB.SI",
    "GAU.SI",
    "GCG.SI",
    "GDM.SI",
    "GED.SI",
    "GET.SI",
    "GEU.SI",
    "GFP.SI",
    "GHM.SI",
    "GI1.SI",
    "GJA.SI",
    "GJL.SI",
    "GMM.SI",
    "GMS.SI",
    "GNU.SI",
    "GP4.SI",
    "GPF.SI",
    "GRN.SI",
    "GRP.SI",
    "GTE.SI",
]


# ── Formatting helpers ────────────────────────────────────────────────────────


def fmt(val, prefix="S$", decimals=2):
    """Return 'prefix + comma-formatted number' or 'N/A' on bad input."""
    try:
        return f"{prefix}{float(val):,.{decimals}f}"
    except (TypeError, ValueError):
        return str(val) if val else "N/A"


def fmt_large(val, prefix="S$"):
    """Format large dollar values with B/M/K suffix (e.g S$4.48B, S$182.5M)."""
    try:
        v = float(val)
        if abs(v) >= 1e12:
            return f"{prefix}{v/1e12:.2f}T"
        elif abs(v) >= 1e9:
            return f"{prefix}{v/1e9:.2f}B"
        elif abs(v) >= 1e6:
            return f"{prefix}{v/1e6:.2f}M"
        elif abs(v) >= 1e3:
            return f"{prefix}{v/1e3:.2f}K"
        else:
            return f"{prefix}{v:.2f}"
    except (TypeError, ValueError):
        return "N/A"


def pct(val):
    """Return a coloured ▲/▼ percentage string, or 'N/A'."""
    try:
        v = float(val)
        arrow = "▲" if v >= 0 else "▼"
        color = "\033[92m" if v >= 0 else "\033[91m"
        return f"{color}{arrow} {abs(v):.2f}%\033[0m"
    except (TypeError, ValueError):
        return "N/A"


def section(title):
    w = 60
    print(f"\n{'-' * w}\n  {title.upper()}\n{'-' * w}")


def row(label, value, width=30):
    print(f"  {label:<{width}} {value}")


# ── yfinance helpers ──────────────────────────────────────────────────────────


def _get_ticker(symbol: str) -> yf.Ticker:
    """Create a yfinance Ticker; Singapore symbols need .SI suffix."""
    sym = symbol.upper().strip()
    if not sym.endswith(".SI"):
        sym = f"{sym}.SI"
    return yf.Ticker(sym)


def _yf_financials(ticker: yf.Ticker):
    """
    Retrieve income statement, balance sheet, and cash flow.
    Handles the yfinance API rename (financials→income_stmt, cashflow→cash_flow).
    """

    def _first_df(*attrs):
        for attr in attrs:
            df = getattr(ticker, attr, None)
            if df is not None and isinstance(df, pd.DataFrame) and not df.empty:
                return df
        return None

    inc = _first_df("income_stmt", "financials")
    bal = _first_df("balance_sheet")
    cf = _first_df("cash_flow", "cashflow")
    return inc, bal, cf


def _row(df, *row_names, col: int = 0):
    """
    Safely fetch a scalar from a yfinance financial DataFrame.
    Tries each row name in order; returns None if none match or value is NaN.
    """
    if df is None or df.empty:
        return None
    for name in row_names:
        if name in df.index:
            try:
                val = df.loc[name].iloc[col]
                return float(val) if pd.notna(val) else None
            except (IndexError, TypeError, ValueError):
                pass
    return None


# ── Data-fetch functions ──────────────────────────────────────────────────────


def fetch_quote(ticker: yf.Ticker) -> dict:
    """Return a unified quote dict from fast_info + info."""
    fi = ticker.fast_info
    info = {}
    try:
        info = ticker.info or {}
    except Exception:
        pass

    def fi_get(attr):
        try:
            v = getattr(fi, attr, None)
            return v if v is not None else None
        except Exception:
            return None

    return {
        # Price data
        "lastPrice": fi_get("last_price")
        or info.get("currentPrice")
        or info.get("regularMarketPrice"),
        "previousClose": fi_get("previous_close") or info.get("previousClose"),
        "open": fi_get("open") or info.get("open"),
        "dayHigh": fi_get("day_high") or info.get("dayHigh"),
        "dayLow": fi_get("day_low") or info.get("dayLow"),
        "volume": fi_get("last_volume") or info.get("volume"),
        "avgVolume": fi_get("three_month_average_volume") or info.get("averageVolume"),
        "marketCap": fi_get("market_cap") or info.get("marketCap"),
        "52wHigh": fi_get("year_high") or info.get("fiftyTwoWeekHigh"),
        "52wLow": fi_get("year_low") or info.get("fiftyTwoWeekLow"),
        "ytdChange": fi_get("year_change"),
        "50dAvg": fi_get("fifty_day_average") or info.get("fiftyDayAverage"),
        "200dAvg": fi_get("two_hundred_day_average") or info.get("twoHundredDayAverage"),
        "currency": fi_get("currency") or info.get("currency", "SGD"),
        # Valuation
        "trailingPE": info.get("trailingPE"),
        "forwardPE": info.get("forwardPE"),
        "trailingEps": info.get("trailingEps"),
        "forwardEps": info.get("forwardEps"),
        "pegRatio": info.get("pegRatio"),
        "priceToBook": info.get("priceToBook"),
        "dividendYield": info.get("dividendYield"),
        "beta": info.get("beta"),
        # Identity
        "name": info.get("shortName") or info.get("longName", "—"),
        "exchange": info.get("exchange") or fi_get("exchange") or "SGX",
        "sector": info.get("sector", "—"),
        "industry": info.get("industry", "—"),
        "country": info.get("country", "SG"),
    }


def fetch_historical(ticker: yf.Ticker, period: str = "1mo") -> pd.DataFrame:
    """Fetch OHLCV history for the given period."""
    try:
        df = ticker.history(period=period, auto_adjust=True)
        return df if isinstance(df, pd.DataFrame) else pd.DataFrame()
    except Exception as e:
        print(f"  [Historical error] {e}")
        return pd.DataFrame()


def fetch_corporate_actions(ticker: yf.Ticker) -> dict:
    """Fetch dividends, stock splits, next earnings."""
    result = {"dividends": [], "splits": [], "calendar": {}}
    try:
        acts = ticker.actions
        if acts is not None and not acts.empty:
            div = acts[acts["Dividends"] > 0]["Dividends"]
            spl = acts[acts["Stock Splits"] > 0]["Stock Splits"]
            result["dividends"] = [
                {"date": str(d.date()), "amount": v} for d, v in div.tail(5).items()
            ]
            result["splits"] = [{"date": str(d.date()), "ratio": v} for d, v in spl.tail(3).items()]
    except Exception:
        pass

    try:
        cal = ticker.calendar or {}
        earnings_dates = cal.get("Earnings Date", [])
        result["calendar"] = {
            "next_earnings": str(earnings_dates[0]) if earnings_dates else "—",
            "ex_dividend_date": str(cal.get("Ex-Dividend Date", "—")),
            "dividend_date": str(cal.get("Dividend Date", "—")),
        }
    except Exception:
        pass

    return result


# ═══════════════════════════════════════════════════════════════════════════════
# SCAN 1 — DARVAS BOX
# ═══════════════════════════════════════════════════════════════════════════════


def compute_darvas_box(df: pd.DataFrame, confirm: int = DARVAS_CONFIRM) -> dict:
    """
    Detect a Darvas Box and classify the current price relative to it.

    KEY DESIGN RULE: Box formation uses ONLY historical bars (all bars except
    the last). The current bar is deliberately excluded so its low cannot
    pull the box bottom down and make a breakdown undetectable.
    """
    if df is None or df.empty:
        return {
            "signal": "INSUFFICIENT_DATA",
            "box_top": None,
            "box_bottom": None,
            "note": "Empty OHLC DataFrame",
        }

    def find_col(df, candidates):
        for c in candidates:
            match = next((col for col in df.columns if c.upper() in col.upper()), None)
            if match:
                return match
        return None

    h_col = find_col(df, ["High", "CH_TRADE_HIGH_PRICE", "DAYHIGH"])
    l_col = find_col(df, ["Low", "CH_TRADE_LOW_PRICE", "DAYLOW"])
    c_col = find_col(df, ["Close", "CH_CLOSING_PRICE", "LAST"])

    if not all([h_col, l_col, c_col]):
        return {
            "signal": "INSUFFICIENT_DATA",
            "box_top": None,
            "box_bottom": None,
            "note": "Could not identify OHLC columns",
        }

    all_highs = pd.to_numeric(df[h_col], errors="coerce").fillna(0).tolist()
    all_lows = pd.to_numeric(df[l_col], errors="coerce").fillna(0).tolist()
    all_closes = pd.to_numeric(df[c_col], errors="coerce").fillna(0).tolist()

    if len(all_closes) < confirm + 5:
        return {
            "signal": "INSUFFICIENT_DATA",
            "box_top": None,
            "box_bottom": None,
            "note": f"Need ≥ {confirm + 5} bars; got {len(all_closes)}",
        }

    current = all_closes[-1]
    highs = all_highs[:-1]
    lows = all_lows[:-1]
    n = len(highs)

    # Step 1: most recent confirmed box top
    box_top_idx = None
    box_top = None
    for i in range(n - confirm - 1, -1, -1):
        candidate = highs[i]
        if candidate == 0:
            continue
        window = highs[i + 1 : i + 1 + confirm]
        if len(window) == confirm and all(h < candidate for h in window):
            box_top_idx = i
            box_top = candidate
            break

    if box_top is None:
        return {
            "signal": "NO_BOX",
            "box_top": None,
            "box_bottom": None,
            "note": "No confirmed box top",
        }

    # Step 2: confirmed box bottom
    segment = lows[box_top_idx:]
    box_bottom = None
    for i in range(len(segment) - confirm):
        candidate = segment[i]
        if candidate == 0:
            continue
        window = segment[i + 1 : i + 1 + confirm]
        if len(window) == confirm and all(l > candidate for l in window):
            box_bottom = candidate
            break

    if box_bottom is None:
        valid = [l for l in segment if l > 0]
        box_bottom = min(valid) if valid else None

    if box_bottom is None:
        return {
            "signal": "NO_BOX",
            "box_top": box_top,
            "box_bottom": None,
            "note": "Could not confirm a box bottom",
        }

    # Step 3: classify current price
    if current > box_top:
        signal = "BREAKOUT_BUY"
    elif current < box_bottom:
        signal = "BREAKDOWN_SELL"
    else:
        signal = "IN_BOX"

    box_range = box_top - box_bottom
    pos_in_box = ((current - box_bottom) / box_range * 100) if box_range else 0
    upside_to_top = ((box_top - current) / current * 100) if current else 0

    return {
        "signal": signal,
        "box_top": round(box_top, 2),
        "box_bottom": round(box_bottom, 2),
        "current_price": round(current, 2),
        "box_range": round(box_range, 2),
        "position_in_box_pct": round(pos_in_box, 1),
        "upside_to_top_pct": round(upside_to_top, 2),
        "confirm_days": confirm,
        "data_points": len(all_closes),
    }


def display_darvas_box(result: dict):
    """Print Darvas Box scan result."""
    section("Darvas Box Scan")
    sig = result.get("signal", "N/A")
    labels = {
        "BREAKOUT_BUY": "\033[92m● BREAKOUT BUY\033[0m  — close above box top",
        "BREAKDOWN_SELL": "\033[91m● BREAKDOWN SELL\033[0m — close below box bottom",
        "IN_BOX": "\033[93m● IN BOX\033[0m        — consolidating",
        "NO_BOX": "No confirmed Darvas box",
        "INSUFFICIENT_DATA": "Insufficient data",
    }
    print(f"\n  Signal: {labels.get(sig, sig)}")
    if result.get("box_top"):
        print()
        row("Box Top", fmt(result["box_top"]))
        row("Box Bottom", fmt(result["box_bottom"]))
        row("Current Price", fmt(result.get("current_price")))
        row("Box Range", fmt(result.get("box_range")))
        row("Position in Box", f"{result.get('position_in_box_pct', 0):.1f}%")
        row("Upside to Top", f"{result.get('upside_to_top_pct', 0):.2f}%")


# ═══════════════════════════════════════════════════════════════════════════════
# SCAN 2 — PIOTROSKI F-SCORE
# ═══════════════════════════════════════════════════════════════════════════════


def compute_piotroski_score(ticker: yf.Ticker, symbol: str = "") -> dict:
    """
    Piotroski F-Score (0–9) from annual financial statements.
    Score ≥ 7 → strong; ≤ 3 → weak.
    """
    inc, bal, cf = _yf_financials(ticker)
    if inc is None or inc.empty:
        return {"symbol": symbol, "error": "No income statement data available"}

    scores = {}
    details = {}

    # Profitability
    net_inc_0 = _row(inc, "Net Income", col=0)
    assets_0 = _row(bal, "Total Assets", col=0)
    roa_0 = (net_inc_0 / assets_0) if (net_inc_0 and assets_0) else None

    net_inc_1 = _row(inc, "Net Income", col=1)
    assets_1 = _row(bal, "Total Assets", col=1)
    roa_1 = (net_inc_1 / assets_1) if (net_inc_1 and assets_1) else None

    scores["F1_ROA_positive"] = 1 if (roa_0 and roa_0 > 0) else 0
    details["ROA_current_%"] = round(roa_0 * 100, 2) if roa_0 else "N/A"

    ocf_0 = _row(cf, "Operating Cash Flow", "Total Cash From Operating Activities", col=0)
    scores["F2_OCF_positive"] = 1 if (ocf_0 and ocf_0 > 0) else 0
    details["OCF_current_$M"] = round(ocf_0 / 1e6, 1) if ocf_0 else "N/A"

    if roa_0 is not None and roa_1 is not None:
        scores["F3_ROA_improving"] = 1 if roa_0 > roa_1 else 0
    else:
        scores["F3_ROA_improving"] = 0
    details["ROA_prev_%"] = round(roa_1 * 100, 2) if roa_1 else "N/A"

    if ocf_0 and assets_0 and roa_0 is not None:
        scores["F4_Accruals"] = 1 if (ocf_0 / assets_0) > roa_0 else 0
    else:
        scores["F4_Accruals"] = 0

    # Leverage & Liquidity
    ltd_0 = _row(bal, "Long Term Debt", col=0) or 0
    ltd_1 = _row(bal, "Long Term Debt", col=1) or 0
    lev_0 = (ltd_0 / assets_0) if assets_0 else None
    lev_1 = (ltd_1 / assets_1) if assets_1 else None
    if lev_0 is not None and lev_1 is not None:
        scores["F5_Leverage_down"] = 1 if lev_0 < lev_1 else 0
    else:
        scores["F5_Leverage_down"] = 0
    details["LTD_ratio_curr_%"] = round(lev_0 * 100, 2) if lev_0 else "N/A"

    ca_0 = _row(bal, "Current Assets", "Total Current Assets", col=0)
    cl_0 = _row(bal, "Current Liabilities", "Total Current Liabilities", col=0)
    ca_1 = _row(bal, "Current Assets", "Total Current Assets", col=1)
    cl_1 = _row(bal, "Current Liabilities", "Total Current Liabilities", col=1)
    cr_0 = (ca_0 / cl_0) if (ca_0 and cl_0) else None
    cr_1 = (ca_1 / cl_1) if (ca_1 and cl_1) else None
    if cr_0 is not None and cr_1 is not None:
        scores["F6_CurrentRatio_up"] = 1 if cr_0 > cr_1 else 0
    else:
        scores["F6_CurrentRatio_up"] = 0
    details["CurrentRatio_curr"] = round(cr_0, 2) if cr_0 else "N/A"

    sh_0 = _row(bal, "Share Issued", col=0)
    sh_1 = _row(bal, "Share Issued", col=1)
    scores["F7_No_dilution"] = (1 if sh_0 <= sh_1 else 0) if (sh_0 and sh_1) else 1

    # Operating Efficiency
    rev_0 = _row(inc, "Total Revenue", col=0)
    gp_0 = _row(inc, "Gross Profit", col=0)
    rev_1 = _row(inc, "Total Revenue", col=1)
    gp_1 = _row(inc, "Gross Profit", col=1)
    gm_0 = (gp_0 / rev_0) if (gp_0 and rev_0) else None
    gm_1 = (gp_1 / rev_1) if (gp_1 and rev_1) else None
    if gm_0 is not None and gm_1 is not None:
        scores["F8_GrossMargin_up"] = 1 if gm_0 > gm_1 else 0
    else:
        scores["F8_GrossMargin_up"] = 0
    details["GrossMargin_curr_%"] = round(gm_0 * 100, 2) if gm_0 else "N/A"

    at_0 = (rev_0 / assets_0) if (rev_0 and assets_0) else None
    at_1 = (rev_1 / assets_1) if (rev_1 and assets_1) else None
    if at_0 is not None and at_1 is not None:
        scores["F9_AssetTurnover_up"] = 1 if at_0 > at_1 else 0
    else:
        scores["F9_AssetTurnover_up"] = 0

    total = sum(scores.values())
    interp = "STRONG" if total >= 7 else "MODERATE" if total >= 4 else "WEAK"
    return {
        "symbol": symbol,
        "f_score": total,
        "interpretation": interp,
        "component_scores": scores,
        "details": details,
    }


def display_piotroski_score(result: dict):
    """Print Piotroski F-Score."""
    section("Piotroski F-Score")
    if "error" in result:
        print(f"  ⚠️  {result['error']}")
        return
    total = result["f_score"]
    color = "\033[92m" if total >= 7 else "\033[93m" if total >= 4 else "\033[91m"
    print(f"\n  Score: {color}{total}/9\033[0m  — {result['interpretation']}")


# ═══════════════════════════════════════════════════════════════════════════════
# SCAN 3 — PEGY RATIO + BREAKOUT + DIVIDEND YIELD
# ═══════════════════════════════════════════════════════════════════════════════


def compute_pegy_and_breakout(ticker: yf.Ticker, symbol: str = "") -> dict:
    """
    PEGY Ratio = PEG ratio + dividend yield (higher yield = attractive entry)
    Breakout Opportunity = Price vs 200-day MA (upside if above, downside if below)
    """
    try:
        ticker.history(period="6mo")
    except Exception:
        return {"symbol": symbol, "error": "Data fetch failed"}

    quote = fetch_quote(ticker)
    cmp = quote.get("lastPrice")
    pe = quote.get("trailingPE")
    div_yield = quote.get("dividendYield") or 0
    ma_200 = quote.get("200dAvg")

    # PEG Ratio (if available from yfinance)
    peg = quote.get("pegRatio")

    # PEGY = PEG + dividend yield (higher = better for value investors)
    pegy = None
    if peg is not None:
        pegy = float(peg) + (float(div_yield) * 100 if div_yield else 0)

    # Breakout opportunity: how far from 200-day MA
    breakout_info = {}
    if cmp and ma_200:
        dist_from_ma = ((cmp - ma_200) / ma_200 * 100) if ma_200 else None
        breakout_info["distance_from_200ma_%"] = round(dist_from_ma, 2) if dist_from_ma else None

        if dist_from_ma and dist_from_ma > 5:
            breakout_info["signal"] = "Above 200-day MA - Uptrend"
        elif dist_from_ma and dist_from_ma < -5:
            breakout_info["signal"] = "Below 200-day MA - Downtrend"
        else:
            breakout_info["signal"] = "Near 200-day MA - Consolidation"

    return {
        "symbol": symbol,
        "cmp": round(cmp, 2) if cmp else None,
        "pe_ratio": round(pe, 2) if pe else None,
        "peg_ratio": round(peg, 2) if peg else None,
        "dividend_yield_%": round(div_yield * 100, 2) if div_yield else None,
        "pegy_adjusted": round(pegy, 2) if pegy else None,
        "ma_200": round(ma_200, 2) if ma_200 else None,
        "breakout": breakout_info,
    }


def display_pegy_and_breakout(result: dict):
    """Print PEGY ratio and breakout analysis."""
    section("PEGY Ratio & Breakout Analysis")
    if "error" in result:
        print(f"  ⚠️  {result['error']}")
        return

    row("Current Market Price (CMP)", fmt(result.get("cmp")))
    row("P/E Ratio", fmt(result.get("pe_ratio"), prefix=""))
    row("PEG Ratio", fmt(result.get("peg_ratio"), prefix=""))
    row("Dividend Yield", f"{result.get('dividend_yield_%', 0):.2f}%")
    row("PEGY Adjusted", f"{result.get('pegy_adjusted', 'N/A')}")
    print()
    row("200-Day MA", fmt(result.get("ma_200")))
    if result.get("breakout"):
        row("Distance from 200MA", f"{result['breakout'].get('distance_from_200ma_%', 0):.2f}%")
        row("Breakout Signal", result["breakout"].get("signal", "N/A"))


# ── Display functions ────────────────────────────────────────────────────────


def display_price_summary(quote: dict):
    """Live price, valuation ratios."""
    section("Live Quote & Valuation")

    ltp = quote.get("lastPrice")
    prev = quote.get("previousClose")
    chg = (ltp - prev) if (ltp and prev) else None
    chg_pct = (chg / prev * 100) if (chg and prev) else None

    row("Company", quote.get("name", "—"))
    row("Exchange", quote.get("exchange", "—"))
    row("Sector", quote.get("sector", "—"))
    row("Currency", quote.get("currency", "SGD"))
    print()
    row("Last Price (CMP)", fmt(ltp))
    row("Prev Close", fmt(prev))
    row("Change", f"{fmt(chg)}  {pct(chg_pct)}")
    row("Day High", fmt(quote.get("dayHigh")))
    row("Day Low", fmt(quote.get("dayLow")))
    print()
    row("52-Week High", fmt(quote.get("52wHigh")))
    row("52-Week Low", fmt(quote.get("52wLow")))
    row("200-Day Avg", fmt(quote.get("200dAvg")))
    print()
    row("Volume", f"{int(quote['volume']):,}" if quote.get("volume") else "N/A")
    row("Market Cap", fmt_large(quote.get("marketCap")))
    print()
    row("Trailing P/E", fmt(quote.get("trailingPE"), prefix=""))
    row("Forward P/E", fmt(quote.get("forwardPE"), prefix=""))
    row("Price/Book", fmt(quote.get("priceToBook"), prefix=""))
    row(
        "Dividend Yield",
        f"{quote['dividendYield']*100:.2f}%" if quote.get("dividendYield") else "N/A",
    )


def display_corporate_actions(actions: dict):
    """Dividends and earnings calendar."""
    section("Corporate Actions")
    cal = actions.get("calendar", {})
    row("Ex-Dividend Date", str(cal.get("ex_dividend_date", "—")))
    row("Dividend Pay Date", str(cal.get("dividend_date", "—")))

    divs = actions.get("dividends", [])
    if divs:
        print("\n  Recent Dividends:")
        for d in reversed(divs):
            print(f"    • {d['date']}  {fmt(d['amount'])} per share")


# ── Main report function ──────────────────────────────────────────────────────


def run(symbol: str, output_format: str = "text", run_scans: bool = False) -> dict:
    """
    Generate a daily stock report for one Singapore equity symbol.

    Args:
        symbol:        Ticker symbol (e.g. 'D05', 'BN4', 'U11') - .SI suffix added automatically
        output_format: 'text' (console) or 'excel' (Excel file)
        run_scans:     Run Darvas Box, Piotroski, PEGY scans

    Returns:
        dict of all fetched and computed data
    """
    symbol = symbol.upper().strip()
    if not symbol.endswith(".SI"):
        symbol = f"{symbol}.SI"

    w = 60
    print(f"\n{'=' * w}")
    print(f"  📊  SINGAPORE STOCK REPORT — {symbol}")
    print(f"  ⏱  Generated: {datetime.now().strftime('%d %b %Y  %H:%M:%S')}")
    print(f"{'=' * w}")

    ticker = _get_ticker(symbol)
    all_data = {"symbol": symbol, "timestamp": datetime.now().isoformat()}

    # Fetch all data
    quote = fetch_quote(ticker)
    fetch_historical(ticker, period="1mo")
    actions = fetch_corporate_actions(ticker)

    # Quantitative scans
    darvas_r = {}
    piotroski_r = {}
    pegy_r = {}

    if run_scans:
        print("  Running quantitative scans …")
        hist_6mo = fetch_historical(ticker, period="6mo")
        darvas_r = compute_darvas_box(hist_6mo)
        piotroski_r = compute_piotroski_score(ticker, symbol)
        pegy_r = compute_pegy_and_breakout(ticker, symbol)

    # Text display
    if output_format == "text":
        display_price_summary(quote)
        display_corporate_actions(actions)

        if run_scans:
            display_darvas_box(darvas_r)
            display_piotroski_score(piotroski_r)
            display_pegy_and_breakout(pegy_r)

        print(f"\n{'=' * w}")
        print("  Data sourced from Yahoo Finance via yfinance")
        print(f"{'=' * w}\n")

    # Prepare data for Excel
    all_data.update(
        {
            "quote": quote,
            "corporate_actions": actions,
            "scans": (
                {
                    "darvas": darvas_r,
                    "piotroski": piotroski_r,
                    "pegy": pegy_r,
                }
                if run_scans
                else {}
            ),
        }
    )

    return all_data


# ── Batch & Excel functions ───────────────────────────────────────────────────


def run_batch(symbols: list = None, output_format: str = "excel", run_scans: bool = False) -> list:
    """
    Run reports for a list of symbols. Defaults to STI-30.
    Saves Excel workbook with summary + per-symbol sheets.
    """
    targets = symbols or STI_30
    print(f"\n{'#' * 60}")
    print(f"  SG BATCH REPORT — {len(targets)} stocks")
    print(f"  Started: {datetime.now().strftime('%d %b %Y  %H:%M:%S')}")
    print(f"{'#' * 60}")

    results, failed = [], []
    for i, sym in enumerate(targets, 1):
        sym_clean = sym.replace(".SI", "")
        print(f"\n[{i:02d}/{len(targets)}] {sym_clean}")
        try:
            data = run(sym, output_format=output_format, run_scans=run_scans)
            results.append(data)
        except Exception as e:
            print(f"  ❌  {sym_clean} failed: {e}")
            failed.append(sym_clean)

    if output_format == "excel" and results:
        _write_excel_report(results, run_scans)

    print(f"\n{'#' * 60}")
    print(f"  Done.  {len(results)} OK, {len(failed)} failed.")
    if failed:
        print(f"  Failed: {', '.join(failed)}")
    print(f"{'#' * 60}\n")
    return results


def _write_excel_report(results: list, include_scans: bool = False):
    """Write comprehensive Excel report with summary sheet + per-stock details."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # ── Create summary sheet ──────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary", 0)

    Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center")

    # Header row
    headers = [
        "Symbol",
        "Company",
        "Sector",
        "CMP (S$)",
        "Change %",
        "P/E",
        "P/B",
        "Div Yield %",
        "52W High",
        "52W Low",
        "Volume",
    ]
    if include_scans:
        headers.extend(
            [
                "Darvas Signal",
                "Darvas Top",
                "Darvas Bottom",
                "Piotroski Score",
                "PEGY Ratio",
                "Breakout Signal",
            ]
        )

    ws_summary.append(headers)
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Data rows
    for data in results:
        sym = data.get("symbol", "").replace(".SI", "")
        q = data.get("quote", {})
        ltp = q.get("lastPrice")
        prev = q.get("previousClose")
        chg_pct = ((ltp - prev) / prev * 100) if (ltp and prev) else None

        row_data = [
            sym,
            q.get("name", ""),
            q.get("sector", ""),
            ltp,
            chg_pct,
            q.get("trailingPE"),
            q.get("priceToBook"),
            (q.get("dividendYield", 0) * 100) if q.get("dividendYield") else None,
            q.get("52wHigh"),
            q.get("52wLow"),
            q.get("volume"),
        ]

        if include_scans:
            scans = data.get("scans", {})
            darv = scans.get("darvas", {})
            piofr = scans.get("piotroski", {})
            pegy = scans.get("pegy", {})

            row_data.extend(
                [
                    darv.get("signal", ""),
                    darv.get("box_top"),
                    darv.get("box_bottom"),
                    piofr.get("f_score"),
                    pegy.get("pegy_adjusted"),
                    pegy.get("breakout", {}).get("signal", ""),
                ]
            )

        ws_summary.append(row_data)

    # Adjust column widths
    for col in ws_summary.columns:
        max_len = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        ws_summary.column_dimensions[col[0].column_letter].width = max_len + 2

    # ── Create per-stock detail sheets ────────────────────────────────────────
    for data in results:
        sym = data.get("symbol", "").replace(".SI", "")
        q = data.get("quote", {})
        scans = data.get("scans", {})

        ws = wb.create_sheet(sym)

        # Title
        ws["A1"] = f"{sym} — {q.get('name', '')}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")

        row_num = 3

        # Quote section
        ws[f"A{row_num}"] = "QUOTE & VALUATION"
        ws[f"A{row_num}"].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f"A{row_num}"].fill = PatternFill(
            start_color="70AD47", end_color="70AD47", fill_type="solid"
        )
        ws.merge_cells(f"A{row_num}:D{row_num}")
        row_num += 1

        quote_fields = [
            ("CMP (S$)", q.get("lastPrice")),
            (
                "Change %",
                (
                    ((q.get("lastPrice") - q.get("previousClose")) / q.get("previousClose") * 100)
                    if (q.get("lastPrice") and q.get("previousClose"))
                    else None
                ),
            ),
            ("52W High", q.get("52wHigh")),
            ("52W Low", q.get("52wLow")),
            ("200-Day MA", q.get("200dAvg")),
            ("P/E Ratio", q.get("trailingPE")),
            ("P/B Ratio", q.get("priceToBook")),
            ("Div Yield %", (q.get("dividendYield", 0) * 100) if q.get("dividendYield") else None),
            ("Market Cap", q.get("marketCap")),
            ("Volume", q.get("volume")),
        ]

        for label, value in quote_fields:
            ws[f"A{row_num}"] = label
            ws[f"B{row_num}"] = value
            row_num += 1

        row_num += 1

        # Scans section
        if include_scans:
            ws[f"A{row_num}"] = "TECHNICAL & FUNDAMENTAL SCANS"
            ws[f"A{row_num}"].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f"A{row_num}"].fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            ws.merge_cells(f"A{row_num}:D{row_num}")
            row_num += 2

            # Darvas
            darv = scans.get("darvas", {})
            ws[f"A{row_num}"] = "Darvas Box"
            ws[f"A{row_num}"].font = Font(bold=True, size=11)
            row_num += 1
            darvas_fields = [
                ("Signal", darv.get("signal")),
                ("Box Top", darv.get("box_top")),
                ("Box Bottom", darv.get("box_bottom")),
                ("Position in Box %", darv.get("position_in_box_pct")),
                ("Upside to Top %", darv.get("upside_to_top_pct")),
            ]
            for label, value in darvas_fields:
                ws[f"A{row_num}"] = label
                ws[f"B{row_num}"] = value
                row_num += 1

            row_num += 1

            # Piotroski
            piofr = scans.get("piotroski", {})
            ws[f"A{row_num}"] = "Piotroski F-Score"
            ws[f"A{row_num}"].font = Font(bold=True, size=11)
            row_num += 1
            ws[f"A{row_num}"] = "Score (0-9)"
            ws[f"B{row_num}"] = piofr.get("f_score")
            row_num += 1
            ws[f"A{row_num}"] = "Interpretation"
            ws[f"B{row_num}"] = piofr.get("interpretation")
            row_num += 1
            row_num += 1

            # PEGY
            pegy = scans.get("pegy", {})
            ws[f"A{row_num}"] = "PEGY Ratio & Breakout"
            ws[f"A{row_num}"].font = Font(bold=True, size=11)
            row_num += 1
            pegy_fields = [
                ("PEG Ratio", pegy.get("peg_ratio")),
                ("PEGY Adjusted", pegy.get("pegy_adjusted")),
                ("Dividend Yield %", pegy.get("dividend_yield_%")),
                ("Breakout Signal", pegy.get("breakout", {}).get("signal")),
                ("Distance from 200MA %", pegy.get("breakout", {}).get("distance_from_200ma_%")),
            ]
            for label, value in pegy_fields:
                ws[f"A{row_num}"] = label
                ws[f"B{row_num}"] = value
                row_num += 1

        # Adjust columns
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    # Save workbook
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = DOWNLOAD_DIR / f"sg_stock_report_{timestamp}.xlsx"
    wb.save(out)
    print(f"\n  📊  Excel report → {out}")


def run_scans_only(symbol: str) -> dict:
    """Run only the three quantitative scans."""
    symbol = symbol.upper().strip()
    if not symbol.endswith(".SI"):
        symbol = f"{symbol}.SI"

    print(f"\n{'=' * 60}\n  🔍  SCANS — {symbol}\n{'=' * 60}")
    ticker = _get_ticker(symbol)
    hist_6mo = fetch_historical(ticker, period="6mo")
    darv = compute_darvas_box(hist_6mo)
    piotr = compute_piotroski_score(ticker, symbol)
    pegy = compute_pegy_and_breakout(ticker, symbol)

    display_darvas_box(darv)
    display_piotroski_score(piotr)
    display_pegy_and_breakout(pegy)

    return {"symbol": symbol, "darvas": darv, "piotroski": piotr, "pegy": pegy}


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Daily Singapore stock report with Darvas Box, Piotroski, and PEGY scans."
    )
    parser.add_argument("symbol", nargs="?", help="SGX ticker symbol (e.g. D05, BN4, U11)")
    parser.add_argument(
        "--scans",
        action="store_true",
        default=False,
        help="Run Darvas Box, Piotroski F-Score, PEGY scans",
    )
    parser.add_argument("--output", choices=["text", "excel"], default="text")
    parser.add_argument(
        "--sti30", action="store_true", default=False, help="Batch report for all 30 STI components"
    )
    parser.add_argument(
        "--sti-top15",
        action="store_true",
        default=False,
        help="Batch report for top 15 STI stocks by market cap",
    )
    parser.add_argument(
        "--sgx-all",
        action="store_true",
        default=False,
        help="Batch report for ALL SGX-listed stocks (~175)",
    )

    args = parser.parse_args()

    if args.sgx_all:
        run_batch(symbols=SGX_ALL, output_format=args.output, run_scans=args.scans)
    elif args.sti30:
        run_batch(symbols=STI_30, output_format=args.output, run_scans=args.scans)
    elif args.sti_top15:
        run_batch(symbols=STI_TOP_15, output_format=args.output, run_scans=args.scans)
    else:
        if not args.symbol:
            print("No symbol given — defaulting to D05 (DBS).")
            args.symbol = "D05"
        run(args.symbol, output_format=args.output, run_scans=args.scans)
