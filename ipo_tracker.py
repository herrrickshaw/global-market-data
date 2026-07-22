# ipo_tracker.py
# ===============
# Subroutine to discover new IPO / fresh listings, download all available
# OHLC history, enrich with financials, run applicable screeners, and
# maintain a live IPO tracking Excel workbook.
#
# HOW NEW LISTINGS ARE DISCOVERED
# ────────────────────────────────
# Strategy: NSE bhavcopy diff
#   Download today's bhavcopy (all EQ symbols trading today).
#   Download the bhavcopy from N days ago (typically 90 days).
#   new_listings = today_symbols - past_symbols
#   This catches IPOs, SME listings, and rights issues that received a fresh symbol.
#
# For 90-day lookback: identifies ~182 new listings (Jun 2026 data).
# For 30-day lookback: finer resolution — catches only very recent IPOs.
#
# SCREENER APPLICABILITY BY LISTING AGE
# ────────────────────────────────────────
#   < 30 days  : Basic metrics only (LTP, listing gain, market cap, PE)
#   30–90 days : Darvas Box starts (needs ≥ 35 bars)
#   3–6 months : Darvas + Magic Formula (if pre-IPO financials in yfinance)
#   6–12 months: + Golden Cross (needs 200 bars ≈ 8 months)
#   > 12 months: + Bull Cartel (needs 5 quarterly results)
#   > 24 months: + Piotroski, Coffee Can (needs 2 years annual data)
#
# 5-YEAR DATA HANDLING
# ─────────────────────
# New stocks don't have 5 years of history.
# The cache stores "max" period (from first trading day onwards).
# Incremental daily updates grow the cache automatically.
# When a stock turns 5 years old, its cache will contain 5 years of data.
#
# USAGE
# ─────
#   python ipo_tracker.py                    # run full IPO scan
#   python ipo_tracker.py --days 30          # only IPOs from last 30 days
#   python ipo_tracker.py --days 365         # full year of new listings
#   python ipo_tracker.py --workers 8
#
# Install:
#   pip install yfinance pandas openpyxl "nse[local]" nsepython

import argparse
import sys
import time
import warnings
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")

try:
    import yfinance as yf

    _YF_OK = True
except ImportError:
    _YF_OK = False
    sys.exit("❌  pip install yfinance")

# ── Constants ─────────────────────────────────────────────────────────────────

OUT_DIR = Path("./ipo_results")
OUT_DIR.mkdir(exist_ok=True)

LOOKBACK_DAYS = 90  # default: find IPOs from last 90 days
MAX_WORKERS = 8
BATCH_SIZE = 50  # smaller batches for IPO stocks (many are illiquid)
SLEEP_BETWEEN = 2.0

# Screener age gates (minimum trading days since listing)
GATE_DARVAS = 35  # needs confirmed box — at least 5 weeks
GATE_MAGIC_FORMULA = 60  # needs financial statements — ~3 months
GATE_GOLDEN_CROSS = 200  # needs 200 DMA — ~8 months
GATE_BULL_CARTEL = 252  # needs 5 quarterly results — ~1 year
GATE_PIOTROSKI = 504  # needs 2 years annual data
GATE_COFFEE_CAN = 756  # needs 3+ years for meaningful CAGR

# ── Disclaimer ────────────────────────────────────────────────────────────────

DISCLAIMER = (
    "⚠️  IPO TRACKER DISCLAIMER: New listings carry higher risk than established stocks. "
    "Limited price history means technical indicators are unreliable. "
    "Financial data may be from DRHP (pre-IPO projections) rather than audited results. "
    "Lock-in periods may affect liquidity. NOT investment advice."
)


# ══════════════════════════════════════════════════════════════════════════════
# 1. DISCOVER NEW LISTINGS
# ══════════════════════════════════════════════════════════════════════════════


def discover_new_listings(lookback_days: int = LOOKBACK_DAYS, verbose: bool = True) -> list[str]:
    """
    Find new NSE EQ listings by comparing today's bhavcopy with N days ago.

    Method: bhavcopy diff
      new_listings = symbols_today - symbols_{lookback_days}_ago

    This catches:
      - Mainboard IPOs (large cap)
      - SME IPOs (BSE SME / NSE Emerge)
      - Re-listings after suspension
      - Rights issue shares receiving new symbols

    Returns list of NSE ticker symbols (new listings only).
    """
    try:
        from nse import NSE

        cache_dir = Path.home() / "Downloads" / "market_cache"
    except ImportError:
        print("  ⚠️  nse library not available — pip install 'nse[local]'")
        return []

    today = datetime.today()
    syms_today = set()
    syms_past = set()

    with NSE(download_folder=str(cache_dir), server=False) as nse_lib:
        # Today's symbols
        for offset in range(7):
            d = today - timedelta(days=offset)
            try:
                result = nse_lib.equityBhavcopy(d)
                if hasattr(result, "exists") and result.exists():
                    df = pd.read_csv(result)
                    if "SctySrs" in df.columns and "TckrSymb" in df.columns:
                        syms_today = set(
                            df[df["SctySrs"] == "EQ"]["TckrSymb"].dropna().str.strip().tolist()
                        )
                        if verbose:
                            print(f"  Today ({d.date()}): {len(syms_today)} EQ symbols")
                        break
            except Exception:
                continue

        # Past symbols (lookback_days ago)
        for offset in range(lookback_days - 5, lookback_days + 10):
            d = today - timedelta(days=offset)
            try:
                result = nse_lib.equityBhavcopy(d)
                if hasattr(result, "exists") and result.exists():
                    df = pd.read_csv(result)
                    if "SctySrs" in df.columns and "TckrSymb" in df.columns:
                        syms_past = set(
                            df[df["SctySrs"] == "EQ"]["TckrSymb"].dropna().str.strip().tolist()
                        )
                        if verbose:
                            print(
                                f"  {lookback_days}d ago ({d.date()}): "
                                f"{len(syms_past)} EQ symbols"
                            )
                        break
            except Exception:
                continue

    if not syms_today or not syms_past:
        print("  ⚠️  Could not fetch bhavcopy for diff — returning empty list")
        return []

    new_listings = sorted(syms_today - syms_past)
    if verbose:
        print(f"\n  NEW LISTINGS in last {lookback_days} days: {len(new_listings)}")
        if new_listings:
            print(f"  Sample: {new_listings[:15]}")
    return new_listings


# ══════════════════════════════════════════════════════════════════════════════
# 2. DOWNLOAD & CACHE OHLC FOR NEW LISTINGS
# ══════════════════════════════════════════════════════════════════════════════


def download_ipo_ohlc(symbols: list[str], verbose: bool = True) -> dict[str, pd.DataFrame]:
    """
    Download maximum available OHLC history for new listings.
    Uses period='max' to get all data from the first trading day onwards.

    New stocks have limited history — we store everything available
    and the cache grows daily via incremental updates.

    For stocks with < 1 year of data, screener gates prevent inappropriate
    technical signals from firing.
    """
    if not _YF_OK:
        return {}

    # Try cache first for already-known stocks
    result: dict[str, pd.DataFrame] = {}
    needs_download = []

    try:
        from market_data_cache import MarketCache

        cache = MarketCache(verbose=False)
        for sym in symbols:
            cached = cache.get_ohlc(f"{sym}.NS")
            if cached is not None and not cached.empty:
                result[sym] = cached
            else:
                needs_download.append(sym)
    except ImportError:
        needs_download = symbols

    if not needs_download:
        if verbose:
            print(f"  All {len(result)} IPO stocks loaded from cache")
        return result

    if verbose:
        print(f"  Downloading {len(needs_download)} new stocks (period=max) …")

    tickers = [f"{s}.NS" for s in needs_download]
    batches = [tickers[i : i + BATCH_SIZE] for i in range(0, len(tickers), BATCH_SIZE)]

    for idx, batch in enumerate(batches, 1):
        if verbose:
            print(f"    Batch {idx}/{len(batches)} ({len(batch)} tickers) …", end=" ", flush=True)
        for attempt in range(3):
            try:
                # period="max" gives all available history from listing date
                raw = yf.download(
                    batch, period="max", auto_adjust=True, threads=True, progress=False
                )
                if raw.empty:
                    print("empty")
                    break
                if isinstance(raw.columns, pd.MultiIndex):
                    for t in batch:
                        sym = t.replace(".NS", "")
                        try:
                            df = raw.xs(t, axis=1, level=1).dropna(how="all")
                            if not df.empty:
                                result[sym] = df[["Open", "High", "Low", "Close", "Volume"]]
                        except KeyError:
                            pass
                else:
                    sym = batch[0].replace(".NS", "")
                    if not raw.empty:
                        result[sym] = raw[["Open", "High", "Low", "Close", "Volume"]]
                ok = sum(1 for t in batch if t.replace(".NS", "") in result)
                if verbose:
                    print(f"OK ({ok}/{len(batch)})")
                break
            except Exception as e:
                if "Rate" in str(e) or "429" in str(e):
                    wait = 20 * (attempt + 1)
                    if verbose:
                        print(f"\n    Rate limited — waiting {wait}s …", end="")
                    time.sleep(wait)
                else:
                    if verbose:
                        print(f"ERROR — {e}")
                    break

        if idx < len(batches):
            time.sleep(SLEEP_BETWEEN)

    # Store new data in cache
    try:
        import json

        from market_data_cache import META_FILE, OHLC_DIR, MarketCache

        cache = MarketCache(verbose=False)
        meta = json.loads(META_FILE.read_text()) if META_FILE.exists() else {}
        for sym, df in result.items():
            if sym in needs_download:
                path = OHLC_DIR / f"{sym}.NS.parquet"
                df.to_parquet(path, compression="snappy", index=True)
                meta[f"ohlc:{sym}.NS"] = {
                    "rows": len(df),
                    "from": str(df.index.min().date()),
                    "to": str(df.index.max().date()),
                    "updated": datetime.now().isoformat(),
                    "file": str(path),
                    "ipo": True,  # flag as IPO stock
                }
        META_FILE.write_text(json.dumps(meta, indent=2, default=str))
        if verbose:
            print(f"  Cache updated: {len(needs_download)} new IPO stocks saved")
    except Exception as e:
        if verbose:
            print(f"  Cache write: {e}")

    if verbose:
        print(f"  Total IPO stocks with OHLC: {len(result)}")
    return result


# ══════════════════════════════════════════════════════════════════════════════
# 3. ENRICH IPO DATA
# ══════════════════════════════════════════════════════════════════════════════


def enrich_ipo(symbol: str, ohlc_df: pd.DataFrame) -> dict:
    """
    Fetch IPO-specific metadata for one stock.

    Returns:
      symbol, company_name, sector, listing_date, listing_price,
      current_ltp, listing_gain%, all_time_high, all_time_low,
      days_since_listing, trading_bars, market_cap, trailing_pe,
      pe_zone, issue_size_cr, screeners_applicable
    """
    result = {
        "Symbol": symbol,
        "Company_Name": symbol,
        "Sector": "—",
        "Listing_Date": None,
        "Listing_Price_₹": None,
        "Current_LTP_₹": None,
        "Listing_Gain_%": None,
        "ATH_₹": None,
        "ATL_₹": None,
        "Days_Listed": 0,
        "Trading_Bars": 0,
        "MCap_Cr": None,
        "Trailing_PE": None,
        "PE_Zone": "⚪ N/A",
        "Screeners_Available": "",
        "Note": "",
    }

    if ohlc_df is None or ohlc_df.empty:
        result["Note"] = "no_ohlc"
        return result

    # OHLC-derived metrics
    closes = ohlc_df["Close"].astype(float)
    result["Trading_Bars"] = len(closes)
    result["Listing_Date"] = str(ohlc_df.index[0].date())
    result["Listing_Price_₹"] = round(float(closes.iloc[0]), 2)
    result["Current_LTP_₹"] = round(float(closes.iloc[-1]), 2)
    result["Listing_Gain_%"] = round(
        (result["Current_LTP_₹"] - result["Listing_Price_₹"]) / result["Listing_Price_₹"] * 100, 2
    )
    result["ATH_₹"] = round(float(ohlc_df["High"].max()), 2)
    result["ATL_₹"] = round(float(ohlc_df["Low"].min()), 2)

    listing_date = ohlc_df.index[0]
    result["Days_Listed"] = (datetime.today() - listing_date).days

    # yfinance info
    try:
        ticker = yf.Ticker(f"{symbol}.NS")
        info = ticker.info or {}
        name = (info.get("shortName") or info.get("longName") or symbol).strip()
        result["Company_Name"] = name[:50]
        result["Sector"] = info.get("sector") or info.get("industry") or "—"
        mcap = info.get("marketCap", 0) or 0
        result["MCap_Cr"] = round(mcap / 1e7, 1) if mcap else None
        tpe = info.get("trailingPE")
        result["Trailing_PE"] = round(float(tpe), 1) if tpe else None

        # PE zone
        from stock_enricher import _classify_pe

        pe_info = _classify_pe(result["Trailing_PE"], result["Sector"], "IN")
        result["PE_Zone"] = f"{pe_info['emoji']} {pe_info['zone']}"
    except Exception:
        pass

    # Screeners applicable based on trading bars
    bars = result["Trading_Bars"]
    applicable = []
    if bars >= GATE_DARVAS:
        applicable.append("Darvas")
    if bars >= GATE_MAGIC_FORMULA:
        applicable.append("MagicFormula")
    if bars >= GATE_GOLDEN_CROSS:
        applicable.append("GoldenCross")
    if bars >= GATE_BULL_CARTEL:
        applicable.append("BullCartel")
    if bars >= GATE_PIOTROSKI:
        applicable.append("Piotroski")
    if bars >= GATE_COFFEE_CAN:
        applicable.append("CoffeeCan")
    result["Screeners_Available"] = ", ".join(applicable) if applicable else "Price only"

    return result


# ══════════════════════════════════════════════════════════════════════════════
# 4. RUN SCREENERS ON NEW LISTINGS
# ══════════════════════════════════════════════════════════════════════════════


def screen_ipo(symbol: str, ohlc_df: pd.DataFrame, meta: dict) -> dict:
    """
    Run screeners that are applicable based on the stock's listing age.
    Returns signal dict with only the signals that are statistically meaningful.
    """
    bars = meta.get("Trading_Bars", 0)
    signals = {"Symbol": symbol}

    if ohlc_df is None or ohlc_df.empty:
        return signals

    # ── Darvas Box (gate: 35 bars) ────────────────────────────────────────────
    if bars >= GATE_DARVAS:
        try:
            from backtest_screeners import detect_darvas_signals

            darvas_sigs = detect_darvas_signals(ohlc_df)
            if darvas_sigs:
                last = darvas_sigs[-1]
                signals["Darvas_Signal"] = "BREAKOUT_BUY"
                signals["Darvas_BoxTop"] = last.get("box_top")
                signals["Darvas_Date"] = str(last.get("date", "")[:10])
            else:
                # Compute current Darvas position
                from full_indian_market_scan import compute_darvas_box

                dr = compute_darvas_box(ohlc_df)
                signals["Darvas_Signal"] = dr.get("signal", "NO_BOX")
                signals["Box_Top"] = dr.get("box_top")
                signals["Box_Bottom"] = dr.get("box_bottom")
                signals["Upside_%"] = dr.get("upside_to_top_pct")
        except Exception as e:
            signals["Darvas_Signal"] = f"error: {e}"

    # ── Golden Cross (gate: 200 bars) ─────────────────────────────────────────
    if bars >= GATE_GOLDEN_CROSS:
        try:
            from full_indian_market_scan import compute_golden_crossover

            gc = compute_golden_crossover(ohlc_df)
            signals["GC_Signal"] = (
                "GOLDEN_CROSS"
                if gc.get("gc_signal")
                else ("ABOVE_200DMA" if gc.get("dma50_above_200") else "BELOW_200DMA")
            )
            signals["DMA50"] = gc.get("dma50")
            signals["DMA200"] = gc.get("dma200")
        except Exception as e:
            signals["GC_Signal"] = f"error: {e}"

    # ── ML Signal (always applicable if enough bars) ──────────────────────────
    if bars >= 60:
        try:
            from ml_signal_engine import MLSignalEngine

            engine = MLSignalEngine(model_type="ridge")
            ml = engine.predict(symbol, ohlc_df)
            signals["ML_Direction"] = ml.get("direction", "NEUTRAL")
            signals["ML_Pred_Ret%"] = ml.get("predicted_ret%", 0)
            signals["ML_Confidence"] = ml.get("confidence", 0)
        except Exception:
            pass

    # ── Fundamental screeners (gate: 60+ bars for financial data) ─────────────
    if bars >= GATE_MAGIC_FORMULA:
        try:
            ticker = yf.Ticker(f"{symbol}.NS")
            inc = _first_df(ticker, "income_stmt", "financials")
            bal = _first_df(ticker, "balance_sheet")
            info = {}
            try:
                info = ticker.info or {}
            except Exception:
                pass
            mcap = info.get("marketCap", 0) or 0

            if inc is not None:
                # Magic Formula
                ebit = _row(inc, "EBIT", "Operating Income", "Ebit")
                a0 = _row(bal, "Total Assets", col=0)
                cl0 = _row(bal, "Current Liabilities", "Total Current Liabilities", col=0)
                cap = (a0 - (cl0 or 0)) if a0 else None
                td = info.get("totalDebt", 0) or 0
                cash = info.get("totalCash", 0) or 0
                ev = (mcap + td - cash) if mcap else None
                roic = (ebit / cap * 100) if (ebit and cap and cap > 0) else None
                ey = (ebit / ev * 100) if (ebit and ev and ev > 0) else None
                bv = info.get("bookValue")
                signals["MF_ROIC_%"] = round(roic, 1) if roic else None
                signals["MF_EY_%"] = round(ey, 1) if ey else None
                signals["MF_Pass"] = (
                    "PASS"
                    if (roic and roic > 15 and ey and ey > 8 and bv and bv > 0 and mcap / 1e7 > 15)
                    else "FAIL"
                )
        except Exception:
            pass

    if bars >= GATE_PIOTROSKI:
        try:
            from full_indian_market_scan import fundamental_scan

            fs = fundamental_scan(symbol, ".NS")
            signals["Piotroski_Score"] = fs.get("f_score")
            signals["Piotroski_Strong"] = "YES" if fs.get("f_strong") else "NO"
            signals["CoffeeCan"] = "PASS" if fs.get("qualifies_cc") else "FAIL"
            signals["BullCartel"] = "PASS" if fs.get("qualifies_bc") else "FAIL"
        except Exception:
            pass

    return signals


# Shared helpers (see stock_utils.py) — aliased to keep existing call sites.
from stock_utils import first_df as _first_df
from stock_utils import row as _row

# ══════════════════════════════════════════════════════════════════════════════
# 5. GENERATE IPO EXCEL
# ══════════════════════════════════════════════════════════════════════════════


def save_ipo_excel(meta_rows: list[dict], signal_rows: list[dict], lookback_days: int) -> Path:
    """
    Save IPO tracking workbook with multiple sheets.

    Sheets:
      DISCLAIMER         Legal/risk warnings for new listings
      IPO_Overview       All new listings: name, dates, price, gain, PE zone
      Price_Analysis     Listing gain ranking + ATH/ATL
      Darvas_Signals     Darvas breakout / breakdown for qualifying IPOs
      ML_Signals         Ridge regression directional signals
      Fundamental_Scan   Magic Formula + Piotroski for older listings
      By_Sector          Grouped by sector
      Screener_Gates     Explains which screeners apply at each age milestone
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    path = OUT_DIR / f"ipo_tracker_{ts}.xlsx"

    meta_df = pd.DataFrame(meta_rows) if meta_rows else pd.DataFrame()
    signal_df = pd.DataFrame(signal_rows) if signal_rows else pd.DataFrame()

    GATES_GUIDE = pd.DataFrame(
        [
            {
                "Trading_Bars": f"<{GATE_DARVAS}",
                "Days_Approx": "<6 weeks",
                "Available_Screeners": "Price only: LTP, listing gain, PE, ATH/ATL",
            },
            {
                "Trading_Bars": f"{GATE_DARVAS}–{GATE_MAGIC_FORMULA}",
                "Days_Approx": "6wk–3mo",
                "Available_Screeners": "Darvas Box breakout",
            },
            {
                "Trading_Bars": f"{GATE_MAGIC_FORMULA}–{GATE_GOLDEN_CROSS}",
                "Days_Approx": "3–8 mo",
                "Available_Screeners": "+ Magic Formula (if pre-IPO financials available)",
            },
            {
                "Trading_Bars": f"{GATE_GOLDEN_CROSS}–{GATE_BULL_CARTEL}",
                "Days_Approx": "8–12 mo",
                "Available_Screeners": "+ Golden Cross (50/200 DMA)",
            },
            {
                "Trading_Bars": f"{GATE_BULL_CARTEL}–{GATE_PIOTROSKI}",
                "Days_Approx": "1–2 yr",
                "Available_Screeners": "+ Bull Cartel (quarterly growth)",
            },
            {
                "Trading_Bars": f">{GATE_PIOTROSKI}",
                "Days_Approx": ">2 years",
                "Available_Screeners": "+ Piotroski F-Score, Coffee Can (full suite)",
            },
        ]
    )

    with pd.ExcelWriter(path, engine="openpyxl") as w:

        def sheet(df, name, sort_col=None, ascending=False):
            if df is None or df.empty:
                pd.DataFrame({"Note": ["No data"]}).to_excel(w, sheet_name=name, index=False)
                return
            if sort_col and sort_col in df.columns:
                df = df.sort_values(sort_col, ascending=ascending)
            df.to_excel(w, sheet_name=name, index=False)

        # Disclaimer
        pd.DataFrame(
            {
                "DISCLAIMER": [
                    DISCLAIMER,
                    "",
                    "KEY RISKS FOR NEW LISTINGS:",
                    "1. Lock-in period: Promoter shares locked for 6–18 months post-IPO",
                    "2. Limited history: Technical screeners unreliable for < 6 months",
                    "3. Financial data: yfinance may show DRHP projections, not audited results",
                    "4. Liquidity: SME IPOs may have wide bid-ask spreads",
                    "5. Valuation: IPO pricing often incorporates peak cycle assumptions",
                    "",
                    f"This tracker covers new NSE EQ listings from the last {lookback_days} days.",
                    "Screener gates prevent inappropriate signals for immature listings.",
                ]
            }
        ).to_excel(w, sheet_name="DISCLAIMER", index=False)

        # All IPOs overview
        if not meta_df.empty:
            sheet(meta_df, "IPO_Overview", sort_col="Days_Listed", ascending=False)

            # Price analysis: sort by listing gain
            sheet(meta_df, "Price_Analysis", sort_col="Listing_Gain_%", ascending=False)

            # By sector
            if "Sector" in meta_df.columns:
                try:
                    sector_df = (
                        meta_df.groupby("Sector")
                        .agg(
                            Count=("Symbol", "count"),
                            Avg_Listing_Gain=("Listing_Gain_%", "mean"),
                            Avg_PE=("Trailing_PE", "mean"),
                            Total_MCap_Cr=("MCap_Cr", "sum"),
                        )
                        .round(2)
                        .reset_index()
                        .sort_values("Count", ascending=False)
                    )
                    sheet(sector_df, "By_Sector")
                except Exception:
                    pass

        # Signals sheets
        if not signal_df.empty:
            darvas_df = (
                signal_df[signal_df.get("Darvas_Signal", "") != ""].copy()
                if "Darvas_Signal" in signal_df.columns
                else pd.DataFrame()
            )
            sheet(darvas_df, "Darvas_Signals", sort_col="Upside_%", ascending=False)

            ml_df = (
                signal_df[signal_df.get("ML_Direction", "") != ""].copy()
                if "ML_Direction" in signal_df.columns
                else pd.DataFrame()
            )
            sheet(ml_df, "ML_Signals", sort_col="ML_Pred_Ret%", ascending=False)

            fund_df = (
                signal_df[signal_df.get("MF_Pass", "").isin(["PASS", "FAIL"])].copy()
                if "MF_Pass" in signal_df.columns
                else pd.DataFrame()
            )
            sheet(fund_df, "Fundamental_Scan", sort_col="MF_ROIC_%", ascending=False)

        # Screener gates guide
        GATES_GUIDE.to_excel(w, sheet_name="Screener_Gates", index=False)

    print(f"\n  📊  IPO Tracker Excel → {path}")
    return path


# ══════════════════════════════════════════════════════════════════════════════
# 6. INTEGRATION — UPDATE MAIN SCAN EXCEL WITH IPO SHEET
# ══════════════════════════════════════════════════════════════════════════════


def append_ipo_to_scan_excel(scan_excel_path: str, meta_rows: list[dict]) -> None:
    """
    Append an IPO_New_Listings sheet to the main full scan Excel workbook.
    Call this after full_indian_market_scan.py completes.
    """
    if not meta_rows:
        return
    try:
        from openpyxl import load_workbook

        wb = load_workbook(scan_excel_path)
        # Remove existing IPO sheet if it exists
        if "IPO_New_Listings" in wb.sheetnames:
            del wb["IPO_New_Listings"]

        # Create new sheet at the end
        ws = wb.create_sheet("IPO_New_Listings")
        df = pd.DataFrame(meta_rows).sort_values("Days_Listed", ascending=False)

        # Write header
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data
        for row_idx, row_data in enumerate(df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(
                    row=row_idx,
                    column=col_idx,
                    value=None if (isinstance(value, float) and np.isnan(value)) else value,
                )

        wb.save(scan_excel_path)
        print(f"  IPO_New_Listings sheet added to {scan_excel_path}")
    except Exception as e:
        print(f"  Could not append IPO sheet: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# 7. MAIN
# ══════════════════════════════════════════════════════════════════════════════


def main(
    lookback_days: int = LOOKBACK_DAYS, workers: int = MAX_WORKERS, append_to_scan: bool = True
):

    print(f"\n{'#'*65}")
    print("  IPO TRACKER — NSE New Listings")
    print(f"  Started: {datetime.now().strftime('%d %b %Y  %H:%M:%S')}")
    print(f"  Lookback: {lookback_days} days | Workers: {workers}")
    print(f"{'#'*65}\n")
    print(DISCLAIMER + "\n")

    # ── Step 1: Discover new listings ─────────────────────────────────────────
    print("Step 1 — Discovering new NSE EQ listings …")
    new_symbols = discover_new_listings(lookback_days=lookback_days)
    if not new_symbols:
        print("  No new listings found — try increasing --days")
        return

    # ── Step 2: Download OHLC (max available history) ─────────────────────────
    print(f"\nStep 2 — Downloading OHLC for {len(new_symbols)} new stocks …")
    ohlc_map = download_ipo_ohlc(new_symbols)

    # ── Step 3: Enrich + screen (parallel) ────────────────────────────────────
    print(f"\nStep 3 — Enriching + screening {len(ohlc_map)} stocks …")
    meta_rows = []
    signal_rows = []
    done = 0

    def _process(sym):
        df = ohlc_map.get(sym)
        m = enrich_ipo(sym, df)
        s = screen_ipo(sym, df, m)
        s.update(
            {
                "Company_Name": m.get("Company_Name", sym),
                "Days_Listed": m.get("Days_Listed", 0),
                "Current_LTP_₹": m.get("Current_LTP_₹"),
            }
        )
        return m, s

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(_process, sym): sym for sym in ohlc_map}
        for future in as_completed(futures):
            sym = futures[future]
            done += 1
            try:
                m, s = future.result()
                meta_rows.append(m)
                signal_rows.append(s)
            except Exception as e:
                meta_rows.append({"Symbol": sym, "Note": str(e)})
            if done % 25 == 0 or done == len(ohlc_map):
                print(f"    {done}/{len(ohlc_map)} processed")

    # ── Step 4: Save IPO Excel ─────────────────────────────────────────────────
    print("\nStep 4 — Saving IPO tracker …")
    path = save_ipo_excel(meta_rows, signal_rows, lookback_days)

    # ── Step 5: Append to latest main scan Excel ───────────────────────────────
    if append_to_scan:
        import glob

        scans = sorted(glob.glob("indian_full_scan/indian_full_scan_*.xlsx"))
        if scans:
            print(f"\nStep 5 — Appending IPO sheet to {scans[-1]} …")
            append_ipo_to_scan_excel(scans[-1], meta_rows)

    # ── Summary ────────────────────────────────────────────────────────────────
    meta_df = pd.DataFrame(meta_rows)
    print(f"\n{'='*65}")
    print(f"  IPO SCAN COMPLETE — {datetime.now().strftime('%d %b %Y  %H:%M:%S')}")
    print(f"  New listings (last {lookback_days}d): {len(new_symbols)}")
    print(f"  With OHLC data:       {len(ohlc_map)}")
    print(
        f"  Darvas-eligible:      {sum(1 for r in meta_rows if r.get('Trading_Bars',0) >= GATE_DARVAS)}"
    )
    print(
        f"  Full-screener-ready:  {sum(1 for r in meta_rows if r.get('Trading_Bars',0) >= GATE_PIOTROSKI)}"
    )

    if not meta_df.empty and "Listing_Gain_%" in meta_df.columns:
        gains = meta_df["Listing_Gain_%"].dropna()
        print("\n  Price performance since listing:")
        print(f"    Avg gain:  {gains.mean():>+.1f}%")
        print(f"    Median:    {gains.median():>+.1f}%")
        print(f"    Best:      {gains.max():>+.1f}%  ({meta_df.loc[gains.idxmax(),'Symbol']})")
        print(f"    Worst:     {gains.min():>+.1f}%  ({meta_df.loc[gains.idxmin(),'Symbol']})")
        print(f"    % positive:{(gains>0).mean()*100:.0f}%")

    if not meta_df.empty:
        print("\n  Top 10 by listing gain:")
        top = meta_df.nlargest(10, "Listing_Gain_%")
        for _, r in top.iterrows():
            gain = r.get("Listing_Gain_%", 0)
            ltp = r.get("Current_LTP_₹", "—")
            bars = r.get("Trading_Bars", 0)
            name = r.get("Company_Name", r.get("Symbol", ""))[:25]
            pe = r.get("PE_Zone", "⚪ N/A")
            print(
                f"    {r['Symbol']:<12} {name:<27} "
                f"LTP:₹{ltp:>8,.0f}  Gain:{gain:>+7.1f}%  "
                f"Bars:{bars:>4}  {pe}"
            )

    print(f"\n  {DISCLAIMER}")
    print(f"{'='*65}\n")
    return {"meta": meta_rows, "signals": signal_rows, "excel": str(path)}


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Track new NSE IPO / fresh listings with full screener analysis.",
        epilog="⚠️  New listings carry higher risk. NOT investment advice.",
    )
    parser.add_argument(
        "--days",
        type=int,
        default=LOOKBACK_DAYS,
        help=f"Lookback window in days (default {LOOKBACK_DAYS})",
    )
    parser.add_argument(
        "--workers", type=int, default=MAX_WORKERS, help=f"Parallel workers (default {MAX_WORKERS})"
    )
    parser.add_argument(
        "--no-append",
        action="store_true",
        default=False,
        help="Skip appending IPO sheet to main scan Excel",
    )
    args = parser.parse_args()

    main(lookback_days=args.days, workers=args.workers, append_to_scan=not args.no_append)
